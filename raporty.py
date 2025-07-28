import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from utils.queries import generate_sample_data
from utils.charts import create_trend_chart, create_summary_chart
import smtplib
import ssl
from email.message import EmailMessage


# PARAMETRY
start_date = datetime(datetime.today().year, 1, 1)
end_date = datetime.today()
output_path = Path("SaldoDzienny.xlsx")

# FIRMY I WALUTY
companies = {
    "Company A": ["PLN", "EUR"],
    "Company B": ["PLN", "EUR"],
}

# GENEROWANIE DANYCH
all_data = []
for company, currencies in companies.items():
    for currency in currencies:
        df = generate_sample_data(company, currency, start_date, end_date)
        all_data.append(df)

# EXCEL
wb = Workbook()
wb.remove(wb.active)

for df in all_data:
    sheet_name = f"{df['Firma'].iloc[0]} {df['Waluta'].iloc[0]}"
    ws = wb.create_sheet(sheet_name)
    
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws[f"A{r_idx}"] = row.Data
        ws[f"B{r_idx}"] = row.Saldo
        ws[f"B{r_idx}"].number_format = '#,##0.00 zł' if row.Waluta == "PLN" else '#,##0.00 €'

    # NAGŁÓWKI
    ws["A1"] = "Data"
    ws["B1"] = "SumaSaldoKoniec"
    ws["A1"].font = ws["B1"].font = Font(bold=True)

    # AUTOSZEROKOŚĆ KOLUMN
    for col in range(1, 3):
        col_letter = get_column_letter(col)
        max_len = max(len(str(ws[f"{col_letter}{r}"].value)) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[col_letter].width = max_len + 4

    # WYKRESY
    chart_buf_full = create_trend_chart(df, f"{sheet_name} – od początku roku", unit='mln' if df['Waluta'].iloc[0] == 'PLN' else 'tys')
    chart_buf_3m = create_trend_chart(df, f"{sheet_name} – ostatnie 3 miesiące", start_date=end_date - timedelta(days=90), unit='mln' if df['Waluta'].iloc[0] == 'PLN' else 'tys')

    if chart_buf_full:
        img = XLImage(chart_buf_full)
        img.anchor = "D2"
        ws.add_image(img)

    if chart_buf_3m:
        img = XLImage(chart_buf_3m)
        img.anchor = "D30"
        ws.add_image(img)

# PODSUMOWANIE
df_all = pd.concat(all_data, ignore_index=True)
summary_ws = wb.create_sheet("Podsumowanie")


summary_ws["A1"] = "Firma"
summary_ws["B1"] = "Waluta"
summary_ws["C1"] = "Saldo końcowe"
for col in ['A1', 'B1', 'C1']:
    summary_ws[col].font = Font(bold=True)

row = 2
for (firma, waluta), group in df_all.groupby(["Firma", "Waluta"]):
    saldo = group.sort_values("Data", ascending=False)["Saldo"].iloc[0]
    summary_ws[f"A{row}"] = firma
    summary_ws[f"B{row}"] = waluta
    summary_ws[f"C{row}"] = saldo
    summary_ws[f"C{row}"].number_format = '#,##0.00 zł' if waluta == "PLN" else '#,##0.00 €'
    row += 1

# WYKRES
buf_pln = create_summary_chart(df_all[df_all["Waluta"] == "PLN"], "Saldo dzienne PLN", start_date=start_date)
buf_eur = create_summary_chart(df_all[df_all["Waluta"] == "EUR"], "Saldo dzienne EUR", start_date=start_date, unit='tys')

if buf_pln:
    img = XLImage(buf_pln)
    img.anchor = "E5"
    summary_ws.add_image(img)

if buf_eur:
    img = XLImage(buf_eur)
    img.anchor = "E35"
    summary_ws.add_image(img)
    
# AUTOSZEROKOŚĆ KOLUMN
for col in range(1, 4):
    col_letter = get_column_letter(col)
    max_len = max(len(str(summary_ws[f"{col_letter}{row}"].value)) if summary_ws[f"{col_letter}{row}"].value else 0
                  for row in range(1, summary_ws.max_row + 1))
    summary_ws.column_dimensions[col_letter].width = max_len + 4


# ZAPIS
wb.save(output_path)
print(f"Plik zapisany: {output_path}")

# MAIL
email_user = "...@..."
email_password = "..."
email_recipient = ["...@..."]
email_subject = "Saldo z kont bankowych"
email_body = "W załączeniu plik Excel z aktualnymi danymi."

msg = EmailMessage()
msg['From'] = email_user
msg['To'] = email_recipient
msg['Subject'] = email_subject
msg.set_content(email_body)

excel_path = "SaldoDzienny.xlsx"
with open(excel_path, 'rb') as f:
    file_data = f.read()
    file_name = "SaldoDzienny.xlsx"
    
msg.add_attachment(file_data, maintype='application', subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=file_name)

context = ssl.create_default_context()
with smtplib.SMTP('smtp.office365.com', 587) as smtp:
    smtp.ehlo() 
    smtp.starttls(context=context)  
    smtp.ehlo()        
    smtp.login(email_user, email_password)
    smtp.send_message(msg)  

print("Mail wysłany.")